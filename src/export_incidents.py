"""Export labor-impact incidents and linked articles to Excel."""

from __future__ import annotations

import argparse
import json
import logging
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.config import DigestConfig
from src.impact_viz import build_labor_impact_viz
from src.labor_impact_parse import VIZ_MIN_DATE
from src.rank import RankedArticle
from src.summarize import snippet_summary
from src.viz_cache import load_payload_cache, year_cache_dir

logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT = ROOT / "data" / "labor_incidents.xlsx"

HEADERS = (
    "Incident #",
    "Date",
    "Region",
    "Country",
    "US state",
    "Industry",
    "AI incident type",
    "Job type",
    "Article headline",
    "Article URL",
    "Summary",
    "Source",
)

URL_COL = HEADERS.index("Article URL") + 1
SUMMARY_COL = HEADERS.index("Summary") + 1


def _load_snippets_by_url() -> dict[str, str]:
    snippets: dict[str, str] = {}
    years_dir = year_cache_dir()
    if not years_dir.is_dir():
        return snippets
    for path in sorted(years_dir.glob("*.json")):
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError, TypeError):
            logger.warning("Skipping unreadable year cache %s", path)
            continue
        for item in raw.get("candidates") or []:
            url = (item.get("url") or "").strip()
            snippet = (item.get("snippet") or "").strip()
            if url and snippet:
                snippets.setdefault(url, snippet)
    return snippets


def _ranked_for_record(record: dict, snippets_by_url: dict[str, str]) -> RankedArticle:
    url = record.get("url") or ""
    snippet = snippets_by_url.get(url, "")
    return RankedArticle(
        headline=record.get("headline") or "",
        url=url,
        score=0,
        reason="",
        text=snippet,
        snippet=snippet,
        source=record.get("source") or "",
        publisher_country=record.get("country") or "",
        thematic_region=record.get("region") or "",
    )


def load_incident_records(*, refresh: bool = False) -> list[dict]:
    """Load parsed incident rows from cache or live discovery."""
    if not refresh:
        payload = load_payload_cache()
        records = (payload or {}).get("records") or []
        if records:
            logger.info("Loaded %d incidents from viz cache", len(records))
            return records

    from src.thematic_regions import default_geographic_regions

    config = DigestConfig(
        date_from=VIZ_MIN_DATE,
        date_to=date.today(),
        article_count=5,
        skip_dedup=True,
        global_coverage=False,
        geographic_regions=default_geographic_regions(),
    )
    result = build_labor_impact_viz(config)
    logger.info("Discovered %d incidents for export", len(result.records))
    return result.records


def build_export_rows(
    records: list[dict],
    snippets_by_url: dict[str, str] | None = None,
) -> list[dict]:
    """One export row per incident (each incident maps to one source article)."""
    snippets = snippets_by_url if snippets_by_url is not None else _load_snippets_by_url()
    rows: list[dict] = []
    for idx, record in enumerate(records, start=1):
        ranked = _ranked_for_record(record, snippets)
        summary = snippet_summary(ranked)
        rows.append(
            {
                "incident_number": idx,
                "date": record.get("date") or "",
                "region": record.get("region") or "",
                "country": record.get("country") or "",
                "us_state": record.get("us_state") or "",
                "industry_type": record.get("industry_type") or "",
                "ai_incident_type": record.get("ai_incident_type") or "",
                "job_type": record.get("job_type") or "",
                "headline": record.get("headline") or "",
                "url": record.get("url") or "",
                "summary": summary,
                "source": record.get("source") or "",
            }
        )
    return rows


def write_incidents_xlsx(rows: list[dict], dest: Path | BinaryIO) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidents"
    ws.append(list(HEADERS))

    header_font = Font(bold=True)
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=col).font = header_font

    link_font = Font(color="0563C1", underline="single")

    for row in rows:
        values = [
            row["incident_number"],
            row["date"],
            row["region"],
            row["country"],
            row["us_state"],
            row["industry_type"],
            row["ai_incident_type"],
            row["job_type"],
            row["headline"],
            row["url"],
            row["summary"],
            row["source"],
        ]
        ws.append(values)
        row_idx = ws.max_row
        url_cell = ws.cell(row=row_idx, column=URL_COL)
        if row["url"]:
            url_cell.hyperlink = row["url"]
            url_cell.font = link_font
        ws.cell(row=row_idx, column=SUMMARY_COL).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "Incident #": 10,
        "Date": 12,
        "Region": 14,
        "Country": 16,
        "US state": 14,
        "Industry": 22,
        "AI incident type": 22,
        "Job type": 24,
        "Article headline": 48,
        "Article URL": 52,
        "Summary": 64,
        "Source": 20,
    }
    for idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)

    ws.freeze_panes = "A2"
    if isinstance(dest, Path):
        dest.parent.mkdir(parents=True, exist_ok=True)
        wb.save(dest)
    else:
        wb.save(dest)


def export_incidents_to_path(
    output: Path = DEFAULT_OUTPUT,
    *,
    refresh: bool = False,
) -> Path:
    records = load_incident_records(refresh=refresh)
    rows = build_export_rows(records)
    write_incidents_xlsx(rows, output)
    logger.info("Wrote %d incidents to %s", len(rows), output)
    return output


def export_incidents_bytes(*, refresh: bool = False) -> bytes:
    records = load_incident_records(refresh=refresh)
    rows = build_export_rows(records)
    buf = BytesIO()
    write_incidents_xlsx(rows, buf)
    return buf.getvalue()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    parser = argparse.ArgumentParser(
        description="Export labor-impact incidents and article summaries to Excel."
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="Re-run discovery instead of using the viz cache",
    )
    args = parser.parse_args()
    path = export_incidents_to_path(args.output, refresh=args.refresh)
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
